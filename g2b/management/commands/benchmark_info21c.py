"""BC-38: 인포21C 벤치마크 커맨드.

인포21C 15건을 BidCompass 엔진으로 재분석해서 head-to-head 비교표 생성.

사용:
    python manage.py benchmark_info21c
    python manage.py benchmark_info21c --input docs/custom.xlsx
    python manage.py benchmark_info21c --skip-api
"""

from __future__ import annotations

import json
import os
import time
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand

from g2b.services.bid_engine import (
    AValueItems,
    TableType,
    WorkType,
    calc_price_score,
    get_floor_rate,
    select_table,
)
from g2b.services.optimal_bid import (
    TABLE_PARAMS_MAP,
    OptimalBidInput,
    compute_expected_score,
    find_optimal_bid,
    generate_scenarios,
)

# API 엔드포인트 (collect_bid_api_data.py와 동일)
BID_INFO_BASE = "https://apis.data.go.kr/1230000/ad/BidPublicInfoService"
AWARD_INFO_BASE = "https://apis.data.go.kr/1230000/as/ScsbidInfoService"

OUTPUT_PATH = Path(settings.BASE_DIR) / "data" / "collected" / "benchmark_info21c.json"


def _parse_amount(val) -> Optional[int]:
    """xlsx 셀값 → int. 콤마 제거, None/빈값 → None."""
    if val is None:
        return None
    s = str(val).replace(",", "").strip()
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _parse_rate(val) -> Optional[float]:
    """투찰율 셀값 → float. 빈값 → None."""
    if val is None:
        return None
    s = str(val).replace(",", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_info21c_xlsx(path: Path) -> list[dict]:
    """인포21C xlsx → 15건 dict 리스트."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    records = []
    for row in range(3, ws.max_row + 1):
        ntce_no = ws.cell(row=row, column=2).value
        if not ntce_no:
            continue

        records.append({
            "bid_ntce_no": str(ntce_no).strip(),
            "bid_name": str(ws.cell(row=row, column=3).value or ""),
            "base_amount": _parse_amount(ws.cell(row=row, column=7).value),
            "a_value": _parse_amount(ws.cell(row=row, column=8).value),
            "estimated_price": _parse_amount(ws.cell(row=row, column=9).value),
            "net_construction_cost": _parse_amount(ws.cell(row=row, column=10).value),
            "info_rate": _parse_rate(ws.cell(row=row, column=15).value),
            "rank1_amount": _parse_amount(ws.cell(row=row, column=18).value),
        })

    wb.close()
    return records


def _get_ntce_ord(bid_ntce_no: str) -> str:
    """공고번호에서 차수 추출: R26BK01351109-000 → 000, 없으면 '00'."""
    if "-" in bid_ntce_no:
        return bid_ntce_no.split("-")[-1]
    return "00"


def _determine_work_type(bid_name: str) -> WorkType:
    """공사명으로 공사구분 판단 (간이: 전기/통신/소방 키워드 → SPECIALTY)."""
    specialty_keywords = ["전기", "정보통신", "소방", "국가유산"]
    for kw in specialty_keywords:
        if kw in bid_name:
            return WorkType.SPECIALTY
    return WorkType.CONSTRUCTION


def _calc_score_for_bid(bid_amount: int, est_price: int, a_value: int,
                        table_type: TableType) -> float:
    """투찰금액으로 가격점수 계산. a_value >= est 면 2.0."""
    if a_value >= est_price:
        return 2.0
    result = calc_price_score(bid_amount, est_price, a_value, table_type)
    return float(result.score)


def run_single_benchmark(record: dict) -> dict:
    """단일 예정가격 기반 비교 (복수예비가격 없는 경우).

    인포: info_rate로 투찰 → 점수
    BC: 90% base rate로 투찰 → 점수
    """
    est = record["estimated_price"]
    a_value = record["a_value"] or 0
    info_rate = record["info_rate"]
    rank1_amount = record["rank1_amount"]
    work_type = _determine_work_type(record["bid_name"])

    table_type = select_table(est, work_type)
    if table_type == TableType.OUT_OF_SCOPE:
        return {**record, "data_source": "out_of_scope", "error": "100억 이상"}

    # 인포 투찰금액 및 점수
    info_bid = int((est - a_value) * (info_rate / 100) + a_value)
    info_score = _calc_score_for_bid(info_bid, est, a_value, table_type)

    # BC: 90% base rate (최적비율)
    bc_bid = int((est - a_value) * Decimal("0.9") + a_value)
    bc_score = _calc_score_for_bid(bc_bid, est, a_value, table_type)

    # 1순위 점수
    rank1_score = (
        _calc_score_for_bid(rank1_amount, est, a_value, table_type)
        if rank1_amount else None
    )

    return {
        "bid_ntce_no": record["bid_ntce_no"],
        "bid_name": record["bid_name"],
        "data_source": "fallback_single_est",
        "table_type": table_type.value,
        "estimated_price": est,
        "a_value": a_value,
        "info_rate": info_rate,
        "info_bid": info_bid,
        "info_score": round(info_score, 2),
        "bc_rate": 90.0,
        "bc_bid": bc_bid,
        "bc_score": round(bc_score, 2),
        "rank1_bid": rank1_amount,
        "rank1_score": round(rank1_score, 2) if rank1_score else None,
        "improvement": round(bc_score - info_score, 2),
    }


def run_optimal_benchmark(record: dict, prelim_prices: list[int]) -> dict:
    """복수예비가격 기반 full 비교 (find_optimal_bid 사용)."""
    est = record["estimated_price"]
    a_value = record["a_value"] or 0
    info_rate = record["info_rate"]
    rank1_amount = record["rank1_amount"]
    work_type = _determine_work_type(record["bid_name"])

    table_type = select_table(est, work_type)
    if table_type == TableType.OUT_OF_SCOPE:
        return {**record, "data_source": "out_of_scope", "error": "100억 이상"}

    # 인포: 단일 예정가격 기준 (인포는 시나리오 분석을 안 하므로)
    info_bid = int((est - a_value) * (info_rate / 100) + a_value)
    info_score = _calc_score_for_bid(info_bid, est, a_value, table_type)

    # BC: 복수예비가격 기반 최적투찰
    inp = OptimalBidInput(
        preliminary_prices=prelim_prices,
        a_value=a_value,
        table_type=table_type,
        presume_price=est,
        net_construction_cost=record.get("net_construction_cost"),
    )
    result = find_optimal_bid(inp)
    bc_bid = result.recommended_bid
    bc_score_expected = result.expected_score

    # 실제 예정가격 기준 BC 점수도 계산 (head-to-head 비교용)
    bc_score_actual = _calc_score_for_bid(bc_bid, est, a_value, table_type)

    # 1순위 점수
    rank1_score = (
        _calc_score_for_bid(rank1_amount, est, a_value, table_type)
        if rank1_amount else None
    )

    return {
        "bid_ntce_no": record["bid_ntce_no"],
        "bid_name": record["bid_name"],
        "data_source": "existing_optimal",
        "table_type": table_type.value,
        "estimated_price": est,
        "a_value": a_value,
        "n_prelim": len(prelim_prices),
        "info_rate": info_rate,
        "info_bid": info_bid,
        "info_score": round(info_score, 2),
        "bc_bid": bc_bid,
        "bc_score": round(bc_score_actual, 2),
        "bc_expected_score": round(bc_score_expected, 2),
        "bc_min_score": round(result.min_scenario_score, 2),
        "bc_max_score": round(result.max_scenario_score, 2),
        "rank1_bid": rank1_amount,
        "rank1_score": round(rank1_score, 2) if rank1_score else None,
        "improvement": round(bc_score_actual - info_score, 2),
    }


def _fetch_prelim_prices_api(bid_ntce_no: str, bid_ntce_ord: str,
                              api_key: str) -> Optional[list[int]]:
    """API로 복수예비가격 15개 수집. 실패/0건 → None."""
    import httpx

    url = f"{AWARD_INFO_BASE}/getOpengResultListInfoCnstwkPPSSrch"
    params = {
        "ServiceKey": api_key,
        "type": "json",
        "numOfRows": 20,
        "pageNo": 1,
        "inqryDiv": "2",
        "bidNtceNo": bid_ntce_no,
        "bidNtceOrd": bid_ntce_ord,
    }

    try:
        with httpx.Client(timeout=10.0) as client:
            for attempt in range(2):
                resp = client.get(url, params=params)
                if resp.status_code == 429:
                    return None  # rate limit
                if resp.status_code != 200:
                    if attempt == 0:
                        time.sleep(1)
                        continue
                    return None
                break

            data = resp.json()
            body = data.get("response", {}).get("body", {})
            items = body.get("items", [])
            if isinstance(items, dict):
                items = items.get("item", [])
                if isinstance(items, dict):
                    items = [items]

            if not items:
                return None

            prices = []
            for item in items:
                val = item.get("bsisPlnprc")
                if val:
                    p = int(float(str(val).strip()))
                    if p > 0:
                        prices.append(p)

            return prices if len(prices) >= 4 else None

    except Exception:
        return None


def _fetch_prelim_from_db(bid_ntce_no: str) -> Optional[list[int]]:
    """DB에서 복수예비가격 조회."""
    from g2b.models import BidApiPrelimPrice

    qs = (
        BidApiPrelimPrice.objects
        .filter(bid_ntce_no=bid_ntce_no, basis_planned_price__gt=0)
        .values_list("basis_planned_price", flat=True)
    )
    prices = list(qs)
    return prices if len(prices) >= 4 else None


class Command(BaseCommand):
    help = "BC-38: 인포21C 15건 벤치마크 비교"

    def add_arguments(self, parser):
        parser.add_argument(
            "--input", type=str, default=None,
            help="xlsx 파일 경로 (기본: docs/인포나의투찰성향목록.xlsx)",
        )
        parser.add_argument(
            "--skip-api", action="store_true",
            help="API 수집 생략 (DB + fallback만)",
        )

    def handle(self, *args, **options):
        input_path = options["input"]
        if input_path:
            xlsx_path = Path(input_path)
        else:
            xlsx_path = Path(settings.BASE_DIR) / "docs" / "인포나의투찰성향목록.xlsx"

        if not xlsx_path.exists():
            self.stderr.write(self.style.ERROR(f"파일 없음: {xlsx_path}"))
            return

        skip_api = options["skip_api"]

        # API 키 (선택)
        api_key = os.getenv("API_KEY", "")
        if not api_key and not skip_api:
            self.stdout.write(self.style.WARNING(
                "API_KEY 미설정 — API 수집 없이 진행"
            ))
            skip_api = True

        # 1. xlsx 파싱
        records = parse_info21c_xlsx(xlsx_path)
        self.stdout.write(f"입력: {len(records)}건 파싱 완료")

        # 2. 각 건 비교 분석
        results = []
        api_stopped = False
        optimal_count = 0
        single_est_count = 0

        for i, rec in enumerate(records, 1):
            ntce_no = rec["bid_ntce_no"]
            ntce_ord = _get_ntce_ord(ntce_no)
            self.stdout.write(f"  [{i}/{len(records)}] {ntce_no} ", ending="")

            if rec["estimated_price"] is None or rec["estimated_price"] <= 0:
                self.stdout.write(self.style.WARNING("예정가격 없음 — skip"))
                continue

            # DB에서 복수예비가격 조회
            prelim = _fetch_prelim_from_db(ntce_no)
            prelim_source = "existing_optimal" if prelim else None

            # DB에 없으면 API로 수집 시도
            if prelim is None and not skip_api and not api_stopped:
                self.stdout.write("(API) ", ending="")
                prelim = _fetch_prelim_prices_api(ntce_no, ntce_ord, api_key)
                if prelim is None:
                    # R-format 아닌 경우 등
                    pass
                else:
                    prelim_source = "api_optimal"
                    self.stdout.write(f"({len(prelim)}개) ", ending="")
                time.sleep(0.3)

            if prelim and len(prelim) >= 4:
                try:
                    result = run_optimal_benchmark(rec, prelim)
                    result["data_source"] = prelim_source
                    optimal_count += 1
                    self.stdout.write(self.style.SUCCESS(
                        f"optimal — info:{result['info_score']} "
                        f"bc:{result['bc_score']} ({result['improvement']:+.2f})"
                    ))
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f"에러: {e}"))
                    result = run_single_benchmark(rec)
                    single_est_count += 1
            else:
                result = run_single_benchmark(rec)
                single_est_count += 1
                self.stdout.write(
                    f"single — info:{result['info_score']} "
                    f"bc:{result['bc_score']} ({result['improvement']:+.2f})"
                )

            results.append(result)

        # 3. 요약 통계
        bc_wins = sum(1 for r in results if r["improvement"] > 0)
        info_wins = sum(1 for r in results if r["improvement"] < 0)
        draws = sum(1 for r in results if r["improvement"] == 0)
        avg_improvement = (
            sum(r["improvement"] for r in results) / len(results)
            if results else 0
        )

        summary = {
            "bc_wins": bc_wins,
            "info_wins": info_wins,
            "draws": draws,
            "avg_improvement": round(avg_improvement, 2),
        }

        # 4. JSON 저장
        output = {
            "meta": {
                "generated_at": datetime.now().isoformat(),
                "input_file": str(xlsx_path),
                "total": len(results),
                "optimal_count": optimal_count,
                "single_est_count": single_est_count,
            },
            "summary": summary,
            "records": results,
        }

        OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS(
            f"결과 저장: {OUTPUT_PATH}"
        ))
        self.stdout.write(
            f"  BC 우위: {bc_wins} / 동점: {draws} / 인포 우위: {info_wins}"
        )
        self.stdout.write(
            f"  평균 개선: {avg_improvement:+.2f}점"
        )
        self.stdout.write(
            f"  최적분석: {optimal_count}건 / 단일예정가격: {single_est_count}건"
        )
